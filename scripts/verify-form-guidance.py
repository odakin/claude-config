#!/usr/bin/env python3
"""
verify-form-guidance.py — 官製様式の「記入要領 (赤字/青字)」 が提出物に残置していないか検出。

WHY: 官製様式は記入要領・placeholder を **赤字 / 青字**で刷り「提出時に削除」 と指示する
(= office-automation.md#docx-guidance-deletion)。 その削除漏れは値検査を全て素通りし、
人の目だけが gate だった。 2026-07-31 に xlsx 出張様式で実発生 (= 休講欄「日付選択」
「時限」 と会期欄「時」「分」 を赤字のまま**印刷**、 user の目視指摘で発覚)。

WHAT (= 2 層。 どちらも **検出のみ。 一切書き換えない**):

  A. xlsx/xlsm — 条件付き書式の containsText ルールを評価する (= **厳密**)
     様式作成者は「この文字列が残っている = 未記入」 を **条件付き書式そのもの**で宣言して
     いることが多い (= `containsText("日付選択")` → dxf font `9C0006` = Excel 標準
     「悪い」 スタイルの赤)。 ∴ ルールの trigger 文字列がその範囲のセルに実在するか、 を
     見れば「記入要領が残っている」 を**様式作成者の定義どおりに**判定できる。 heuristic
     ではなく様式自身の宣言なので偽陽性が原理的に無い。
     ⚠️ **cell.font.color を見る検査は空振りする** — 赤は cell font ではなく条件付き書式
     由来で、 font.color は黒 (theme 1) のままだから (= 実際に空振りした経路)。

  B. pdf — 描画後の span 色を見る (= **heuristic**、 A の網羅外を拾う保険)
     直接 font 色 / style 継承 / 図形内文字など、 A で見えない経路の色付き文字を拾う。
     PDF は「Excel/Word が最終的に描いた色」 なので継承解決済 (= 同 slug 落とし穴 a/b)。

  A が xlsx 段、 B が PDF 段なので **両方を driver に噛ませる**のが本来の使い方
  (A = 生成前に落とす / B = 生成後に落とす)。

判定と exit code:
  🔴 FAIL (exit 1) = A の trigger 残置 / B の赤・青 span   → 提出前に人間が判断して消す
  🟡 WARN (exit 0) = B のその他の非黒 span                  → 情報提供のみ
  **消す判断は人間**。 本 script は「様式構造の見出し」 と「記入要領」 を区別できない
  (= office-automation.md#docx-guidance-deletion (1) の境界判断は人間の仕事)。

LIMIT (= 何を検出しないか。 「0 件 = 安全」 と読み替えないための宣言):
  - A は **条件付き書式で宣言された placeholder のみ**。 ルールを持たない様式 (実測: 学外者
    様式・謝金様式は CF ルール 0 件) では A は常に 0 件を返す → その様式は B だけが頼り。
  - A は文字列セルのみ照合する (数式セル・数値セルの表示文字列は評価しない)。
  - A は赤系・青系 dxf のルールだけを FAIL にする (= 緑等で「正常」 を示す CF を誤検出しない)。
  - B は色だけを見る ∴ **赤い見出し**を持つ様式では偽陽性になる → `--allow-color` で除外。
    ⚠️ 色は描画で ±1 ずれる (実測: 指定 `0070C0` → PDF span `0070BF`) ので、 `--allow-color`
    の照合は各チャネル ±`ALLOW_TOL` の許容差で行う (= 完全一致にすると除外が空振りする)。
  - 印影・ロゴ等の画像は span ではないので B の対象外 (= 赤い印鑑で FAIL しない)。

USAGE:
  verify-form-guidance.py <file> [<file> ...]        # .xlsx/.xlsm → A、 .pdf → B (拡張子で分岐)
  verify-form-guidance.py --allow-color 0070C0 f.pdf # この色を FAIL にしない
  verify-form-guidance.py --quiet ...                # finding が無ければ何も出さない
  verify-form-guidance.py --selftest                 # 内蔵 self-test (外部 file 不要)
"""

from __future__ import annotations

import sys
from pathlib import Path

FAIL, WARN = "FAIL", "WARN"

# Excel 標準「悪い」 スタイルの font 色 (= 官製様式の記入要領で最頻)。 実測 PDF span も同値。
EXCEL_BAD_RED = (0x9C, 0x00, 0x06)
ALLOW_TOL = 3          # --allow-color の許容差 (= 描画 round-trip の ±1 を吸収)


def _rgb(v):
    """'FF9C0006' / '9C0006' / int / (r,g,b) を (r,g,b) に正規化。 不明は None。"""
    if v is None:
        return None
    if isinstance(v, tuple):
        return v if len(v) == 3 else None
    if isinstance(v, int):
        return ((v >> 16) & 255, (v >> 8) & 255, v & 255)
    s = str(v).strip()
    if len(s) == 8:          # ARGB
        s = s[2:]
    if len(s) != 6:
        return None
    try:
        n = int(s, 16)
    except ValueError:
        return None
    return ((n >> 16) & 255, (n >> 8) & 255, n & 255)


def classify(rgb):
    """色を black / red / blue / other に分類 (= 記入要領は赤か青、 が様式の慣習)。"""
    if rgb is None:
        return "unknown"
    r, g, b = rgb
    if r < 60 and g < 60 and b < 60:
        return "black"
    if r >= 100 and g <= 90 and b <= 90:
        return "red"
    if b >= 100 and r <= 90 and g <= 130:
        return "blue"
    return "other"


# ---------------------------------------------------------------- A: xlsx
def check_xlsx(path):
    """条件付き書式 containsText の trigger 文字列が範囲内に実在するセルを列挙。"""
    try:
        import openpyxl
    except ImportError:
        return [(WARN, "openpyxl 不在 = xlsx 検査 skip (pip install openpyxl)")]

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path)

    out = []
    for ws in wb.worksheets:
        seen = set()                       # 同じルールが 2 回登録されている様式があるので de-dup
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type != "containsText":
                    continue
                text = getattr(rule, "text", None)
                if not text:
                    continue
                dxf = getattr(rule, "dxf", None)
                color = None
                if dxf is not None and dxf.font is not None and dxf.font.color is not None:
                    color = dxf.font.color.rgb
                kind = classify(_rgb(color))
                if kind not in ("red", "blue"):
                    continue               # 緑等 = 「正常」 を示す CF。 記入要領ではない
                for cr in cf.sqref.ranges:
                    for row in ws[str(cr)]:
                        cells = row if isinstance(row, tuple) else (row,)
                        for cell in cells:
                            v = cell.value
                            if not isinstance(v, str) or text not in v:
                                continue
                            key = (ws.title, cell.coordinate, text)
                            if key in seen:
                                continue
                            seen.add(key)
                            out.append((FAIL,
                                        f"[{ws.title}] {cell.coordinate}: 記入要領 "
                                        f"{text!r} が残置 (値 {v!r}, 条件付き書式 "
                                        f"{kind} {color} = 様式作成者が「未記入」 と宣言)"))
    return out


# ---------------------------------------------------------------- B: pdf
def check_pdf(path, allow=()):
    """描画済 PDF の非黒 text span を色で分類して報告。"""
    try:
        import fitz
    except ImportError:
        return [(WARN, "PyMuPDF 不在 = pdf 検査 skip (pip install pymupdf)")]

    allowed = [a for a in (_rgb(x) for x in allow) if a]

    def is_allowed(rgb):
        return rgb is not None and any(
            all(abs(rgb[i] - a[i]) <= ALLOW_TOL for i in range(3)) for a in allowed)

    buckets = {}
    d = fitz.open(path)
    try:
        for pi in range(d.page_count):
            for blk in d[pi].get_text("dict")["blocks"]:
                for line in blk.get("lines", []):
                    for sp in line.get("spans", []):
                        t = (sp.get("text") or "").strip()
                        if not t:
                            continue
                        rgb = _rgb(sp.get("color"))
                        kind = classify(rgb)
                        if kind == "black" or is_allowed(rgb):
                            continue
                        buckets.setdefault((kind, rgb), []).append((pi + 1, t))
    finally:
        d.close()

    out = []
    for (kind, rgb), items in sorted(buckets.items(), key=lambda kv: -len(kv[1])):
        sample = " / ".join(f"p{p}:{t[:20]!r}" for p, t in items[:5])
        lvl = FAIL if kind in ("red", "blue") else WARN
        hint = ("提出前に削除要否を判断 (様式構造の見出しなら残す)" if lvl == FAIL
                else "情報提供のみ")
        out.append((lvl, f"{kind} rgb={rgb} の文字 {len(items)} 件: {sample} … {hint}"))
    return out


# ---------------------------------------------------------------- driver
def check_file(path, allow=()):
    suf = Path(path).suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return check_xlsx(path)
    if suf == ".pdf":
        return check_pdf(path, allow)
    return [(WARN, f"未対応の拡張子 {suf} (対象 = .xlsx/.xlsm/.pdf)")]


# ---------------------------------------------------------------- selftest
def _selftest():
    import tempfile
    results = []

    def rec(name, hits, want_fail):
        got = len([h for h in hits if h[0] == FAIL])
        ok = (got >= 1) if want_fail else (got == 0)
        results.append((name, got, want_fail, ok))

    try:
        import openpyxl
        from openpyxl.formatting.rule import Rule
        from openpyxl.styles.differential import DifferentialStyle
        from openpyxl.styles import Font
    except ImportError:
        print("SKIP: openpyxl 不在 = xlsx selftest 省略")
        openpyxl = None

    if openpyxl:
        def mk(values, color="FF9C0006", sqref="M38:R40", text="日付選択"):
            wb = openpyxl.Workbook()
            ws = wb.active
            dxf = DifferentialStyle(font=Font(color=color))
            ws.conditional_formatting.add(
                sqref, Rule(type="containsText", operator="containsText",
                            text=text, dxf=dxf,
                            formula=[f'NOT(ISERROR(SEARCH("{text}",M38)))']))
            for k, v in values.items():
                ws[k] = v
            f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            f.close()  # Windows: 開いた handle が残ると save() が Permission denied
            wb.save(f.name)
            return f.name

        rec("xlsx clean (記入済)", check_xlsx(mk({"M38": "2026年8月7日（金）"})), False)
        rec("xlsx 残置 (anchor cell)", check_xlsx(mk({"M38": "日付選択"})), True)
        rec("xlsx 残置 (範囲内の非 anchor cell)", check_xlsx(mk({"M40": "日付選択"})), True)
        rec("xlsx 部分一致 (前後に文字)", check_xlsx(mk({"M38": "← 日付選択 ←"})), True)
        rec("xlsx 緑 CF は無視 (選択性)",
            check_xlsx(mk({"M38": "日付選択"}, color="FF00B050")), False)
        rec("xlsx 青 CF も FAIL",
            check_xlsx(mk({"M38": "日付選択"}, color="FF0070C0")), True)
        rec("xlsx 日付値は非該当 (型)",
            check_xlsx(mk({"M38": __import__("datetime").datetime(2026, 8, 7)})), False)

        # --- retroactive fixture: 2026-07-31 の実事故を様式実測値で再現 -------------
        # 1_1【教育職員用】国内出張様式 の実 CF (sqref / trigger / 色) をそのまま置き、
        # 事故当時の残置値を入れる → 6 セル全部が FAIL に出ることを要求する。
        wb = openpyxl.Workbook()
        ws = wb.active
        real = [("M25:X25", "日付選択"), ("M26:X28", "日付選択"),
                ("AC25:AF28", "時"), ("AE25:AF28", "分"),
                ("M38:R40", "日付選択"), ("S38:U40", "時限")]
        for sqref, text in real:
            ws.conditional_formatting.add(
                sqref, Rule(type="containsText", operator="containsText", text=text,
                            dxf=DifferentialStyle(font=Font(color="FF9C0006")),
                            formula=[f'NOT(ISERROR(SEARCH("{text}",A1)))']))
        for k, v in {"M38": "日付選択", "S38": "時限", "M39": "日付選択", "S39": "時限",
                     "M40": "日付選択", "S40": "時限", "AC25": "時", "AE25": "分"}.items():
            ws[k] = v
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        f.close()  # Windows: 開いた handle が残ると save() が Permission denied
        wb.save(f.name)
        hits = check_xlsx(f.name)
        n = len([h for h in hits if h[0] == FAIL])
        results.append(("xlsx 実事故 fixture (2026-07-31、 ≥8 セル)", n, True, n >= 8))

    try:
        import fitz
    except ImportError:
        print("SKIP: PyMuPDF 不在 = pdf selftest 省略")
        fitz = None

    if fitz:
        def mkpdf(items):
            d = fitz.open()
            pg = d.new_page()
            y = 72
            for text, color in items:
                pg.insert_text((72, y), text, fontsize=11, color=color)
                y += 20
            f = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            f.close()  # Windows: 開いた handle が残ると save() が Permission denied
            d.save(f.name)
            d.close()
            return f.name

        black, red, blue, gray = (0, 0, 0), (0.61, 0, 0.02), (0, 0.44, 0.75), (0.5, 0.5, 0.5)
        rec("pdf clean (黒のみ)", check_pdf(mkpdf([("filled value", black)])), False)
        rec("pdf 赤字残置", check_pdf(mkpdf([("date placeholder", red)])), True)
        rec("pdf 青字残置", check_pdf(mkpdf([("guidance note", blue)])), True)
        rec("pdf 灰は WARN 止まり", check_pdf(mkpdf([("watermark", gray)])), False)
        rec("pdf --allow-color で除外",
            check_pdf(mkpdf([("heading", blue)]), allow=["0070C0"]), False)

    ok = True
    for name, got, want, good in results:
        print(f"  {'✅' if good else '❌'} {name}: FAIL {got} 件 "
              f"(期待 {'≥1' if want else '0'})")
        ok &= good
    print("selftest:", "ALL PASS" if ok else "FAIL")
    return 0 if ok else 1


def main(argv):
    args = list(argv)
    if "--selftest" in args or not args:
        return _selftest()
    quiet = "--quiet" in args
    args = [a for a in args if a != "--quiet"]
    allow = []
    while "--allow-color" in args:
        i = args.index("--allow-color")
        allow.append(args[i + 1])
        del args[i:i + 2]

    rc = 0
    for p in args:
        findings = check_file(p, allow)
        fails = [f for f in findings if f[0] == FAIL]
        if findings:
            print(f"── {Path(p).name}")
            for lvl, msg in findings:
                print(f"   {'🔴' if lvl == FAIL else '🟡'} {msg}")
        elif not quiet:
            print(f"── {Path(p).name}: 記入要領 (赤字/青字) の残置なし")
        if fails:
            rc = 1
    return rc


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
