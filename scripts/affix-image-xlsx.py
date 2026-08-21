#!/usr/bin/env python3
"""Place an image (seal / signature) into an .xlsx via Excel.app — without destroying the file.

Why not openpyxl
----------------
``openpyxl``'s ``add_image`` + ``wb.save()`` looks like the obvious route, but on a
real office template it silently damages two things (measured 2026-07-27 on a
university travel form):

===================  ==================  ==================
part                 before              after openpyxl
===================  ==================  ==================
xl/drawings/*.vml    6                   6      (kept)
xl/comments*.xml     6                   **12** (duplicated)
formula cache        populated           **all None**
===================  ==================  ==================

Duplicated comment parts make Excel report the workbook as damaged, and the wiped
formula cache means every ``=IF(other!A1,...)`` reads back empty until someone
opens the file in Excel again. Driving Excel itself has neither effect: the same
measurement gives 6 / 6 / cache intact.

So: this script asks Excel to insert the picture, then *verifies* that nothing else
moved. See ``office-automation.md#xlsx-image-via-excel``.

Pre-granted staging dir (2026-08-21)
------------------------------------
Excel is App-Sandboxed: the first time it has to *write* into a folder it pops its
own "ファイル アクセスを許可" dialog (once per new folder, unclickable when driven
remotely). Both the workbook and the image are therefore copied into the Office App
Group container (``~/Library/Group Containers/UBF8T346G9.Office/claude-office-staging/
<unique>/``), Excel edits the *copy* there, the post-condition checks run on the
copy, and only then is it copied back over the original (atomic ``os.replace``).
A failed run leaves the original byte-identical and keeps the staged dir for
diagnosis. ``--no-stage`` / ``CLAUDE_OFFICE_STAGING=0`` = in-place (old behaviour).
Lib = ``scripts/lib/office_staging.py``, doc =
``office-automation.md#office-pregranted-staging-dir``.

Usage
-----
    affix-image-xlsx.py BOOK.xlsx --sheet 依頼書 --cell AH11 --image seal.png
    affix-image-xlsx.py BOOK.xlsx --sheet 依頼書 --cell AH11 --image seal.png \
        --size 34 --dx -10 --dy 2
    affix-image-xlsx.py BOOK.xlsx --sheet 依頼書 --cell AH11 --image seal.png --dry-run

``--cell`` anchors the image's top-left corner to that cell's top-left corner;
``--dx``/``--dy`` shift it in points (positive = right / down).

Safety
------
* Refuses to run when Excel already has the target workbook open (avoids clobbering
  unsaved edits).
* Writes a ``.bak`` next to the file unless ``--no-backup`` (in-place mode; with
  staging the original is never touched until the checks pass).
* Post-condition assertions abort with exit 2 and restore the backup (in-place) /
  keep the original untouched (staging) if the drawing inventory or formula cache
  changed in a way other than "one new image".
* ``--dry-run`` reports the computed position without touching the workbook.
"""
from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
import zipfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib"))
from office_staging import Stage  # noqa: E402


def inventory(path: str) -> dict:
    """Count the fragile parts of an xlsx (drawings / vml / comments / media)."""
    names = zipfile.ZipFile(path).namelist()
    return {
        "drawing": len([n for n in names if "drawings/drawing" in n]),
        "vml": len([n for n in names if n.endswith(".vml")]),
        "comments": len([n for n in names if "comments" in n and n.endswith(".xml")]),
        "media": len([n for n in names if "/media/" in n]),
    }


def formula_cache_sample(path: str, limit: int = 12) -> list:
    """Sample cached values of cross-sheet formulas, to catch a cache wipe."""
    try:
        import openpyxl
    except ImportError:
        return []
    live = openpyxl.load_workbook(path)
    cached = openpyxl.load_workbook(path, data_only=True)
    out = []
    for name in live.sheetnames:
        ws, wc = live[name], cached[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    out.append((name, cell.coordinate, wc[cell.coordinate].value))
                    if len(out) >= limit:
                        return out
    return out


def run_osascript(script: str) -> str:
    proc = subprocess.run(["osascript", "-"], input=script, capture_output=True, text=True)
    if proc.returncode != 0:
        sys.exit(f"osascript failed:\n{proc.stderr.strip()}")
    return proc.stdout.strip()


def excel_has_open(path: str) -> bool:
    """True when Excel is running with this workbook open."""
    running = run_osascript(
        'tell application "System Events" to return (name of processes) contains "Microsoft Excel"'
    )
    if running != "true":
        return False
    names = run_osascript(
        'tell application "Microsoft Excel" to if running then return name of every workbook'
    )
    return os.path.basename(path) in names


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("book")
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--cell", required=True, help="anchor cell, e.g. AH11")
    ap.add_argument("--image", required=True)
    ap.add_argument("--size", type=float, default=34.0, help="width=height in points (default 34)")
    ap.add_argument("--dx", type=float, default=0.0, help="horizontal offset in points")
    ap.add_argument("--dy", type=float, default=0.0, help="vertical offset in points")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-backup", action="store_true")
    ap.add_argument("--no-stage", action="store_true",
                    help="drive Excel on the original in place (skip the pre-granted staging dir)")
    args = ap.parse_args()

    book = os.path.abspath(args.book)
    image = os.path.abspath(args.image)
    for p in (book, image):
        if not os.path.exists(p):
            sys.exit(f"not found: {p}")

    if excel_has_open(book):
        sys.exit(f"Excel already has {os.path.basename(book)} open — close it first (unsaved edits would be lost).")

    if args.no_stage:
        os.environ["CLAUDE_OFFICE_STAGING"] = "0"

    before_inv = inventory(book)
    before_cache = formula_cache_sample(book)

    with Stage(book, image) as st:
        sbook, simage = st.paths
        if st.active:
            print(f"staging: {st.dir} (pre-granted, no sandbox dialog)", file=sys.stderr)

        if args.dry_run:
            pos = run_osascript(f'''
tell application "Microsoft Excel"
  open POSIX file "{sbook}"
  delay 1
  tell workbook 1
    set r to range "{args.cell}" of worksheet "{args.sheet}"
    set L to (left position of r)
    set T to (top of r)
  end tell
  close workbook 1 saving no
  return (L as text) & "," & (T as text)
end tell''')
            left, top = (float(v) for v in pos.split(","))
            print(f"[dry-run] {args.sheet}!{args.cell} -> left={left + args.dx} top={top + args.dy} size={args.size}")
            print(f"[dry-run] inventory: {before_inv}")
            return

        backup = book + ".bak"
        if not args.no_backup and not st.active:
            shutil.copy2(book, backup)   # in-place mode only; staging never touches the original early

        run_osascript(f'''
tell application "Microsoft Excel"
  open POSIX file "{sbook}"
  delay 1
  tell workbook 1
    tell worksheet "{args.sheet}"
      set r to range "{args.cell}"
      set L to (left position of r) + {args.dx}
      set T to (top of r) + {args.dy}
      make new picture at it with properties {{file name:(POSIX file "{simage}"), left position:L, top:T, width:{args.size}, height:{args.size}}}
    end tell
    save
  end tell
  close workbook 1 saving no
end tell''')

        after_inv = inventory(sbook)
        after_cache = formula_cache_sample(sbook)

        problems = []
        if after_inv["media"] != before_inv["media"] + 1:
            problems.append(f"media {before_inv['media']} -> {after_inv['media']} (expected +1)")
        if after_inv["vml"] != before_inv["vml"]:
            problems.append(f"vml {before_inv['vml']} -> {after_inv['vml']} (expected unchanged)")
        if after_inv["comments"] != before_inv["comments"]:
            problems.append(f"comments {before_inv['comments']} -> {after_inv['comments']} (expected unchanged)")
        lost = [(s, c) for (s, c, v), (_, _, v2) in zip(before_cache, after_cache) if v is not None and v2 is None]
        if lost:
            problems.append(f"formula cache lost at {lost[:5]}")

        if problems:
            if st.active:
                restored = f" (original untouched; staged copy kept at {st.dir})"
            elif not args.no_backup and os.path.exists(backup):
                shutil.move(backup, book)
                restored = " (restored from backup)"
            else:
                restored = " (NO BACKUP — file left as-is)"
            sys.exit("post-condition failed:\n  " + "\n  ".join(problems) + restored)

        if st.active:
            st.copy_back(sbook, book)
        elif not args.no_backup and os.path.exists(backup):
            os.remove(backup)
    print(f"ok: {os.path.basename(image)} -> {args.sheet}!{args.cell}  inventory {before_inv} -> {after_inv}")


if __name__ == "__main__":
    main()
