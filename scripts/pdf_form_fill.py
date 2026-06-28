#!/usr/bin/env python3
"""雛形 PDF への直接印字エンジン (= office-automation.md #pdf-prefill-direct の汎用実装)。

行政・学術様式の「標題・押印マーク等が drawing の xlsx」 を openpyxl で編集すると drawing が
全消失する (#openpyxl-destroys-drawings)。 紙提出だけが必要な場合の最速安全経路 =
雛形 xlsx を Excel で PDF 化 (drawing は render 済) → その PDF に本エンジンで値を印字。

組み込み済みの安全装置 (= 2026-06-11 の印刷事故 3 連 RCA を全て機械化):
  - 座標は label 語の bbox から導出 (hardcode しない、 雛形改訂に頑健)
  - 文字照合は全て NFKC (= CJK 互換字形の false negative 回避, #pdf-text-match-nfkc)
  - `=TODAY()` 由来の `#+` overflow は redact で除去 (矩形 shrink で隣接巻き添え防止)
  - フォントは実 file 埋め込み + subset (組み込み "japan" は glyph 不描画 renderer あり)
  - 検証内蔵 (全挿入値の存在 + `##` 残存 + ページ数)
  - 印刷用に 600dpi ラスタ版も生成 (= subset font の printer RIP 化け対策, #print-raster-pdf)

使い方 (library): 様式ごとの driver script が import して使う。
    from pdf_form_fill import build_document
    build_document(template_pdf="form.pdf",
                   page_contains=["銀行振込口座"], page_not_contains=[],
                   items=[{"anchor": "所属:", "occurrence": 0, "dx": 8, "text": "〇〇学科"},
                          {"anchor": "殿", "align": "right", "dx": -4, "text": "山田 太郎"}],
                   out_base="out/dir/書類名")
  → 書類名_filled.pdf (確認用) + 書類名_raster.pdf (印刷用) を生成、 検証 fail は例外。

item の仕様:
  anchor      : ページ上の label 語 (NFKC 一致、 get_text("words") の 1 語)
  occurrence  : 同語が複数あるとき何番目か (y→x 順、 default 0)
  dx, dy      : anchor の右端 (align=left) / 左端 (align=right) からの offset (pt)
  align       : "left" (= anchor の右に置く、 default) / "right" (= text 右端を anchor 左端に合わせる)
  text        : 印字する文字列。 "\n" 区切りで複数行 (行送りは fontsize*1.45)
  size        : fontsize (default 9)
  verify      : False にすると検証対象から外す (= "✓" 等、 重複しうる短い記号用)
  type        : "check" で anchor (= "□..." 等の checkbox 語) の □ 内に ✓ をベクター描画
                (= font の ✓ glyph 有無に非依存)。 text 不要、 検証・二重印字 guard 対象外。
"""

from __future__ import annotations

import re
import shutil
import subprocess
import unicodedata
from pathlib import Path

import fitz

# overlay フォント候補 (= macOS の既知 path)。 ⚠️ 先頭 = 游ゴシック Regular (= Office 同梱)
# = Excel が吐く雛形 PDF の既定日本語フォント。 雛形と overlay でフォントが違うと太さ・字形が
# 不揃いになる (#pdf-prefill-font-match。 Arial Unicode は太く雛形と不揃いになった実害が origin)。
# 非 macOS (= Linux 等) ではこれらは存在しないので pick_font が fontconfig (fc-match) で
# Noto Sans CJK 等に解決し、 それも無ければ build_document(font=...) を要求する。
FONT_CANDIDATES = [
    "/Applications/Microsoft Excel.app/Contents/Resources/DFonts/YuGothR.ttc",
    "/Applications/Microsoft Word.app/Contents/Resources/DFonts/YuGothR.ttc",
    "/Applications/Microsoft PowerPoint.app/Contents/Resources/DFonts/YuGothR.ttc",
    "/Library/Fonts/Arial Unicode.ttf",          # fallback (= 太め、 雛形と不揃いになる)
    "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
]
# 雛形 PDF の埋込フォント基底名 (= subset prefix 除去後) → 揃える system font file。
KNOWN_TEMPLATE_FONTS = {
    "YuGothic": [
        "/Applications/Microsoft Excel.app/Contents/Resources/DFonts/YuGothR.ttc",
        "/Applications/Microsoft Word.app/Contents/Resources/DFonts/YuGothR.ttc",
    ],
    "Hiragino": ["/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc"],
}
XLSX_TO_PDF = Path(__file__).resolve().parent / "xlsx-to-pdf.sh"

# 各種ダッシュ/ハイフンを ASCII '-' に畳む (= 埋込フォント subset がハイフン '-' を
# 抽出時 U+2010/U+2011 等に round-trip するため、 検証の substring 照合が空振りする。
# #pdf-text-match-nfkc の dash 拡張)。
_DASH_MAP = {ord(c): "-" for c in "‐‑‒–—―−﹣－"}


def nfkc(s: str) -> str:
    return unicodedata.normalize("NFKC", s)


def flat(s: str) -> str:
    """NFKC + ダッシュ正規化 (= 照合専用、 描画は変えない)。"""
    return nfkc(s).translate(_DASH_MAP)


def _first_existing(paths):
    return next((p for p in paths if Path(p).exists()), None)


def _has_cjk(path: str) -> bool:
    """font file が CJK glyph を持つか (= 「日」 で判定)。 fc-match は family 不在時に
    非 CJK のデフォルト font を返すため、 採用前に必ず CJK 被覆を検証する (= 豆腐防止)。"""
    try:
        return bool(fitz.Font(fontfile=path).has_glyph(ord("日")))
    except Exception:
        return False


def _fc_match(family: str):
    """非 macOS: fontconfig (fc-match) で family 名 → 実 font file を解決 (= OS 非依存の CJK 解決)。
    ⚠️ fc-match は不在 family でもデフォルト font を返すので、 CJK 被覆を検証してから返す。"""
    if not shutil.which("fc-match"):
        return None
    try:
        out = subprocess.run(["fc-match", "-f", "%{file}", family],
                             capture_output=True, text=True, timeout=5)
        p = out.stdout.strip()
        return p if p and Path(p).exists() and _has_cjk(p) else None
    except Exception:
        return None


# 非 macOS で CJK glyph を持つ font を family 名で探す順 (= fontconfig 経由)。
_CJK_FAMILIES = ("Yu Gothic", "Noto Sans CJK JP", "Noto Sans JP", "Hiragino Sans", "IPAGothic")


def pick_font(template_pdf=None) -> str:
    """overlay フォントを選ぶ。 template_pdf を渡すと雛形の埋込フォントに合わせる
    (#pdf-prefill-font-match)。 macOS は既知 path、 非 macOS は fontconfig で CJK font を解決。
    どれも無ければ FileNotFoundError (= build_document(font=...) で明示指定を要求)。"""
    if template_pdf is not None:
        try:
            doc = fitz.open(str(template_pdf))
            names = [f[3].split("+")[-1] for pg in doc for f in pg.get_fonts(full=True)]
            body = [n for n in names if "Bold" not in n]   # 値の本文は Regular に揃える
            for key, files in KNOWN_TEMPLATE_FONTS.items():
                if any(key in n for n in body):
                    hit = _first_existing(files) or _fc_match("Yu Gothic" if key == "YuGothic" else key)
                    if hit:
                        return hit
        except Exception:
            pass
    hit = _first_existing(FONT_CANDIDATES)          # macOS 既知 path (= 高速)
    if hit:
        return hit
    for fam in _CJK_FAMILIES:                        # 非 macOS (Linux 等): fontconfig で名前解決
        fc = _fc_match(fam)
        if fc:
            return fc
    raise FileNotFoundError(
        "CJK glyph を持つ font が見つかりません。 macOS は Office (游ゴシック) / Arial Unicode、 "
        "非 macOS は fontconfig に Noto Sans CJK 等を入れるか、 build_document(font='/path/to.ttf') "
        "で明示指定してください。")


def ensure_template_pdf(template_xlsx: Path) -> Path:
    """xlsx → PDF (Excel 経由、 drawing render 済)。 PDF が新しければ再生成しない。"""
    pdf = template_xlsx.with_suffix(".pdf")
    if not pdf.exists() or pdf.stat().st_mtime < template_xlsx.stat().st_mtime:
        subprocess.run(["zsh", str(XLSX_TO_PDF), str(template_xlsx)],
                       check=True, capture_output=True)
    if not pdf.exists():
        raise RuntimeError(f"雛形 PDF 生成失敗: {template_xlsx} (Excel が必要)")
    return pdf


def assert_formula_cache_intact(workbook_path, sentinel_cells) -> None:
    """openpyxl save 後の周辺 cell の formula cache 消失を検出 (= driver の冒頭 reflex)。

    openpyxl `wb.save()` は cell formula の AST は保持するが cached value を再計算せず
    drop する。 後段で `data_only=True` で読む driver は値を空欄として読み、 PDF に空欄
    印字 / `#REF!` などの silent failure に至る (= cell value 一致検証では検出できない)。
    本 helper は openpyxl 編集後の xlsx を読み込む直後に呼んで、 不変 sentinel cell の
    cache が空判定なら driver を loud fail させる (= sentinel guard pattern を engine
    helper として hoist、 各 driver で同じ inline コードを書かなくて良くする)。

    workbook_path  : 検査対象 xlsx の path (= openpyxl 編集を経た出力 xlsx)
    sentinel_cells : list of (sheet_name, cell_ref) tuples。 各 cell は要件 3 つ:
                     (a) **編集 script が触らない位置** (= 不変)
                     (b) **別 sheet 参照 formula で cache 依存** (= cache 喪失 → None)
                     (c) **元 file で formula が valid に解決済** (= 非 None が期待値)
                     driver の最終出力 PDF に直接影響しない位置である必要は **無い**
                     (= cache 依存性が本質、 「出力に使わない」 は要件ではない)。

    Raises SystemExit (with helpful message) if any sentinel cache is None。
    sheet 自体が存在しなければその sentinel は silent skip (= driver 側の責任で、
    bootstrap 期の sheet 追加と本 helper の同時 wire を許容)。

    See: claude-config/conventions/office-automation.md §openpyxl-clears-formula-cache
    修復: Excel.app で xlsx を open+save 1-pass (= 全 formula を再計算して cache を埋め直す)。
    """
    from openpyxl import load_workbook
    wb_cached = load_workbook(workbook_path, data_only=True)
    failures = []
    for sheet_name, cell_ref in sentinel_cells:
        if sheet_name not in wb_cached.sheetnames:
            continue  # silent skip = caller の責任 (= sheet 不在の form もあるため)
        val = wb_cached[sheet_name][cell_ref].value
        if val is None:
            failures.append(f"{sheet_name}!{cell_ref}")
    if failures:
        raise SystemExit(
            f"⚠️ formula cache 空: {failures}。 openpyxl save が cache を破壊した可能性。\n"
            f"   Excel.app で {workbook_path} を open+save 1-pass 走らせて再計算を強制 → 再実行してください。\n"
            f"   詳細: office-automation.md §openpyxl-clears-formula-cache"
        )


def find_page(doc: "fitz.Document", contains: list, not_contains: list = ()) -> int:
    """内容特徴語でページを特定 (= ページ番号 hardcode 禁止)。"""
    for i, page in enumerate(doc):
        t = nfkc(page.get_text())
        if all(c in t for c in contains) and not any(c in t for c in not_contains):
            return i
    raise LookupError(f"該当ページなし: contains={contains}")


def word_rect(page, label: str, occurrence: int = 0) -> "fitz.Rect":
    hits = [fitz.Rect(w[:4]) for w in page.get_text("words") if nfkc(w[4]) == nfkc(label)]
    hits.sort(key=lambda r: (round(r.y0), r.x0))
    if len(hits) <= occurrence:
        raise LookupError(f"anchor 語「{label}」(#{occurrence}) が見つからない (雛形改訂?)")
    return hits[occurrence]


def _rect_area(r: "fitz.Rect") -> float:
    return abs(r.width * r.height)


def find_data_cell_for_label(page, label_text: str, direction: str = None,
                             occurrence: int = 0) -> "fitz.Rect":
    """label を含む cell に隣接する text-empty data cell の rect を返す
    (= label cell vs data cell の判別罠を構造的に消す helper)。

    日本の行政・学術様式の標準 layout は「label cell + 隣接 data cell」 の 2 cell 構造で、
    label cell に書かれた「○○欄/(自筆にて記載)」 を reflex で記入欄と取り違える事故が起きる
    (= [`pdf-cell-label-vs-data-disambiguation`](office-automation.md#pdf-cell-label-vs-data-disambiguation))。
    本 helper は label cell を機械で identify し、 隣接する text ゼロの真 data cell を返す
    ことで、 driver の reflex 判断を不要化する (= helper を使う限り label cell に overlay
    する事故は起きない)。

    direction  : 'right' / 'below' / 'above' / 'left' / None (= 全方向探索、 right > below
                 > above > left の優先順)。 日本様式の最頻 pattern は右側または下側だが、
                 form 設計に依存するため呼出し側で direction を pin できる
    occurrence : 同 label が複数欄で出る form (例: 「氏名」 が複数箇所) で何番目かを選ぶ
                 (= search_for 結果の y→x 順 indexing、 default 0)

    Algorithm (= 2-tier、 PDF cell の描かれ方の variant を吸収):
      tier 1 = filled-rect approach: label を contain する rect 群を **area 降順** で
               列挙し (= 「outer block cell = 真の label cell unit」 を先に試す、 inner-
               tight cell は drawing 構造の artifact で偽 hit の origin)、 各々で direction
               方向の adjacent filled rect (= border share ∧ text ゼロ) を探索。 Excel-
               derived PDF で cell が opaque な filled rect として描かれている場合に hit
      tier 2 = border-segment approach: 各 containing cell について direction 方向の最近接
               border line segment (= 細い rect = horizontal/vertical line) から data cell
               の rect を構築 (= label cell の y range を borrow + 右/下/上/左の line までを span)。
               Excel-derived PDF で cell が 4 本の thin border + transparent fill として
               描かれている場合 (= 推薦書 PDF の実観測 pattern) に hit

    Returns: fitz.Rect (= 真の data cell)。 caller は insert_image / insert_text の rect
             として直接渡せる。 cell border に喰い込みたくなければ margin を別途引く。

    Raises:
      LookupError if (a) label 不在 / (b) containing rect ゼロ (= image-only PDF で
                     drawings 無し、 本 helper 非対応) / (c) 全 containing cell で
                     direction 方向の empty data cell 候補ゼロ。 (c) は debug 用に
                     候補列挙を message に含める

    See: office-automation.md#pdf-cell-label-vs-data-disambiguation
    """
    PRIORITY = ("right", "below", "above", "left")
    dirs = (direction,) if direction else PRIORITY
    for d in dirs:
        if d not in PRIORITY:
            raise ValueError(f"direction must be one of {PRIORITY} or None, got {d!r}")

    hits = page.search_for(nfkc(label_text))
    if len(hits) <= occurrence:
        raise LookupError(
            f"label「{label_text}」(#{occurrence}) が見つからない "
            f"(雛形改訂? hit={len(hits)})"
        )
    label_bbox = hits[occurrence]

    drawings = page.get_drawings()
    all_rects = [d["rect"] for d in drawings if d.get("rect")]
    if not all_rects:
        raise LookupError(
            f"page に drawings 無し (= image-only PDF か、 雛形が cell 構造を持たない)。 "
            f"本 helper は cell border を持つ雛形 PDF 専用 (label={label_text!r})"
        )

    AREA_FLOOR = 50.0
    TOL = 2.0

    # containing cells = label bbox を完全に含む rect を **area 降順** (= 大きい
    # 「label cell unit」 から先に探索)。 nest 構造では outer block が真の label cell で、
    # inner-tight cell は「文字 tight 囲み」 でしかない (= drawing 構造の artifact、 その
    # 右隣探索は outer cell 内部を hit して偽 data cell を返す失敗 mode の origin)。
    # 大きい cell の隣接で hit しなかった時のみ小さい cell に fallback (= robustness)。
    containing = sorted(
        [r for r in all_rects if r.contains(label_bbox) and _rect_area(r) >= AREA_FLOOR],
        key=_rect_area, reverse=True,
    )
    if not containing:
        raise LookupError(
            f"label「{label_text}」 を contain する rect 無し (label_bbox={label_bbox})"
        )

    def is_h_line(r):
        return r.height < 1.5 and r.width > 5

    def is_v_line(r):
        return r.width < 1.5 and r.height > 5

    debug_attempts = []

    def adjacent_filled(cell, d):
        """tier 1: cell に direction d で隣接する filled rect で text ゼロのものを返す。
        無ければ None。"""
        for r in all_rects:
            if r == cell or _rect_area(r) < AREA_FLOOR:
                continue
            if d == "right":
                shares = abs(r.x0 - cell.x1) < TOL and r.y0 < cell.y1 - TOL and r.y1 > cell.y0 + TOL
            elif d == "left":
                shares = abs(r.x1 - cell.x0) < TOL and r.y0 < cell.y1 - TOL and r.y1 > cell.y0 + TOL
            elif d == "below":
                shares = abs(r.y0 - cell.y1) < TOL and r.x0 < cell.x1 - TOL and r.x1 > cell.x0 + TOL
            elif d == "above":
                shares = abs(r.y1 - cell.y0) < TOL and r.x0 < cell.x1 - TOL and r.x1 > cell.x0 + TOL
            else:
                continue
            if shares and not page.get_text(clip=r).strip():
                return r
        return None

    def adjacent_from_borders(cell, d):
        """tier 2: cell の direction d で最近接 border line から data cell rect を構築。
        無ければ None。"""
        if d == "right":
            verts = sorted(
                [r for r in all_rects if is_v_line(r) and r.x0 > cell.x1
                 and r.y0 < cell.y1 + TOL and r.y1 > cell.y0 - TOL],
                key=lambda r: r.x0,
            )
            if not verts:
                return None
            cand = fitz.Rect(cell.x1, cell.y0, verts[0].x0, cell.y1)
        elif d == "left":
            verts = sorted(
                [r for r in all_rects if is_v_line(r) and r.x1 < cell.x0
                 and r.y0 < cell.y1 + TOL and r.y1 > cell.y0 - TOL],
                key=lambda r: -r.x1,
            )
            if not verts:
                return None
            cand = fitz.Rect(verts[0].x1, cell.y0, cell.x0, cell.y1)
        elif d == "below":
            horizs = sorted(
                [r for r in all_rects if is_h_line(r) and r.y0 > cell.y1
                 and r.x0 < cell.x1 + TOL and r.x1 > cell.x0 - TOL],
                key=lambda r: r.y0,
            )
            if not horizs:
                return None
            cand = fitz.Rect(cell.x0, cell.y1, cell.x1, horizs[0].y0)
        elif d == "above":
            horizs = sorted(
                [r for r in all_rects if is_h_line(r) and r.y1 < cell.y0
                 and r.x0 < cell.x1 + TOL and r.x1 > cell.x0 - TOL],
                key=lambda r: -r.y1,
            )
            if not horizs:
                return None
            cand = fitz.Rect(cell.x0, horizs[0].y1, cell.x1, cell.y0)
        else:
            return None
        if _rect_area(cand) < AREA_FLOOR:
            return None
        if page.get_text(clip=cand).strip():
            return None
        return cand

    for cell in containing:
        for d in dirs:
            r = adjacent_filled(cell, d)
            if r:
                return r
            debug_attempts.append(f"tier1 {d} of cell {cell}: miss")
            r = adjacent_from_borders(cell, d)
            if r:
                return r
            debug_attempts.append(f"tier2 {d} of cell {cell}: miss")

    raise LookupError(
        f"label「{label_text}」 の隣接 data cell 候補ゼロ "
        f"(direction={direction or 'any'}, occurrence={occurrence})。\n"
        f"  containing cells: {[str(c) for c in containing]}\n"
        f"  attempts: {debug_attempts[:8]}{'...' if len(debug_attempts) > 8 else ''}"
    )


def boost_signature_alpha(src_png, alpha_boost: float = 3.0,
                          force_black: bool = True) -> "Path":
    """透過 PNG 署名の濃度を boost した temp PNG path を返す
    (= [`signature-image-overlay-density`](office-automation.md#signature-image-overlay-density)
    の経路非依存 helper)。

    自筆署名の scan PNG は antialias edge の灰色 RGB + 中間 alpha の pixel が多いため、
    どの renderer (fitz / Word / Pages / LibreOffice) も元より淡く composite する。
    本 helper は PIL で 2 段階前処理 (alpha × N で半透明 pixel を不透明寄り + RGB を pure
    black に固定で antialias 灰色 edge を真っ黒へ) した temp PNG を返す。

    alpha_boost : alpha channel × N (上限 255 で clip)。 1.0=原寸、 1.8=~80% 濃く、
                  3.0=ほぼ全 alpha を 255 化。 線形に効くので段階的 iteration 推奨
    force_black : True なら alpha > 0 pixel の RGB を (0, 0, 0) に固定 (= antialias の灰色
                  edge を真っ黒へ)。 False なら手書き濃淡を温存 (= 「digital 加工バレ」 を
                  警戒する窓口や法的契約向け)

    Returns: Path to temp PNG。 ⚠️ caller 責任で .unlink(missing_ok=True) で cleanup。

    Raises:
      ImportError if PIL + numpy 不在 (= `pip3 install Pillow numpy` を message に案内)
      ValueError if alpha_boost < 0

    See: office-automation.md#signature-image-overlay-density
    """
    if alpha_boost < 0:
        raise ValueError(f"alpha_boost must be >= 0, got {alpha_boost}")
    try:
        from PIL import Image
        import numpy as np
    except ImportError as e:
        raise ImportError(
            "boost_signature_alpha requires Pillow + numpy. "
            "Install with: pip3 install Pillow numpy"
        ) from e
    import tempfile
    import os

    img = Image.open(src_png).convert("RGBA")
    r, g, b, a = img.split()
    a_arr = np.array(a, dtype=np.float32) * float(alpha_boost)
    a_arr = np.clip(a_arr, 0, 255).astype(np.uint8)
    a_new = Image.fromarray(a_arr)
    if force_black:
        zero = Image.new("L", img.size, 0)
        out = Image.merge("RGBA", (zero, zero, zero, a_new))
    else:
        out = Image.merge("RGBA", (r, g, b, a_new))
    fd, tmp_name = tempfile.mkstemp(suffix=".png", prefix="sig_boosted_")
    os.close(fd)
    tmp = Path(tmp_name)
    out.save(tmp)
    return tmp


def _draw_check(page, r: "fitz.Rect") -> None:
    """anchor (= "□普通" 等の checkbox 語) の □ 内に ✓ をベクター描画。
    font の ✓ glyph 有無に依存せず確実に印字する (= □ は語頭の全角1字、 r.x0 が □ の左端)。"""
    bw = r.y1 - r.y0
    x0 = r.x0 + 1.0
    p1 = fitz.Point(x0 + 1.0, r.y0 + bw * 0.55)
    p2 = fitz.Point(x0 + bw * 0.42, r.y1 - 1.5)
    p3 = fitz.Point(x0 + bw * 0.95, r.y0 + 1.5)
    page.draw_line(p1, p2, width=1.1, color=(0, 0, 0))
    page.draw_line(p2, p3, width=1.1, color=(0, 0, 0))


def redact_hash_runs(page) -> int:
    """`####` 等 (= =TODAY() の列幅 overflow、 個数は出力時の列幅依存) を除去。"""
    rects = [fitz.Rect(w[:4]) for w in page.get_text("words") if re.fullmatch(r"#+", w[4])]
    for r in rects:
        page.add_redact_annot(fitz.Rect(r.x0 + 2, r.y0 + 1, r.x1, r.y1 - 1))
    if rects:
        page.apply_redactions()
    return len(rects)


def redact_words(page, words: list) -> int:
    """数式由来の不要表示 (= 例: 空参照の "0") を語単位で除去。 矩形は shrink。"""
    rects = [fitz.Rect(w[:4]) for w in page.get_text("words") if w[4] in words]
    for r in rects:
        page.add_redact_annot(fitz.Rect(r.x0 + 0.5, r.y0 + 0.5, r.x1 - 0.5, r.y1 - 0.5))
    if rects:
        page.apply_redactions()
    return len(rects)


def build_document(template_pdf, page_contains, items, out_base,
                   page_not_contains=(), drop_words=(), dpi=600,
                   font=None, check_double_print=True, assert_present=()) -> dict:
    """1 書類 (= 1 ページ) を生成。 return = {"filled": path, "raster": path}。

    font=None なら雛形 PDF の埋込フォントに自動マッチ (#pdf-prefill-font-match)。
    check_double_print=True で「雛形に既に存在する値を再印字 = 二重印字」 を検出して
    例外 (#pdf-prefill-template-prefilled)。 申請者欄等が雛形に prefill 済の様式で、
    その値を誤って item に入れると重なる事故を loud fail させる (= 黙って二重刷りを防ぐ)。
    雛形が legitimately 同値を持つ item は `allow_preexisting: True` で個別 opt-out。
    assert_present = 雛形に prefill 済で item には入れない値 (= 申請者ブロック等) のうち、
    出力に存在することを sanity-check したいもの。 別財源で雛形を差し替えて prefill 値が
    欠落した時に loud fail させる (= check_double_print の対称: 再印字を止める一方、 こちらは
    「在るべき prefill が消えていないか」 を検証。 #pdf-prefill-template-prefilled)。"""
    font = font or pick_font(template_pdf)
    src = fitz.open(str(template_pdf))
    pno = find_page(src, list(page_contains), list(page_not_contains))
    doc = fitz.open()
    doc.insert_pdf(src, from_page=pno, to_page=pno)
    page = doc[0]

    # --- 二重印字 guard (= 雛形に既に値がある欄を item に入れていないか、 印字前に検出) ---
    if check_double_print:
        base_t = flat(page.get_text()).replace(" ", "").replace("\n", "")
        clashes = sorted({
            str(it["text"]) for it in items
            if it.get("type") != "check"
            and it.get("verify", True) and not it.get("allow_preexisting", False)
            and len(str(it["text"]).strip()) >= 2
            and flat(str(it["text"]).replace("\n", "")).replace(" ", "") in base_t
        })
        if clashes:
            raise AssertionError(
                f"二重印字の恐れ ({out_base}): 次の値は雛形 PDF に既に存在する "
                f"(= item から外すか allow_preexisting:True、 #pdf-prefill-template-prefilled): {clashes}")

    redact_hash_runs(page)
    if drop_words:
        redact_words(page, list(drop_words))

    overlay_font = fitz.Font(fontfile=font)   # = 右寄せ印字の文字幅計測用 (= fitz.Font.text_length)
    for it in items:
        r = word_rect(page, it["anchor"], it.get("occurrence", 0))
        if it.get("type") == "check":
            _draw_check(page, r)
            continue
        size = it.get("size", 9)
        F = dict(fontname="JPF", fontfile=font, fontsize=size)
        lines = str(it["text"]).split("\n")
        if it.get("align", "left") == "right":
            width = max(overlay_font.text_length(ln, fontsize=size) for ln in lines)
            x = r.x0 - width + it.get("dx", -4)
        else:
            x = r.x1 + it.get("dx", 6)
        y = r.y1 + it.get("dy", 0)
        for ln in lines:
            page.insert_text((x, y), ln, **F)
            y += size * 1.45

    doc.subset_fonts()
    out_filled = Path(f"{out_base}_filled.pdf")
    out_filled.parent.mkdir(parents=True, exist_ok=True)
    doc.save(out_filled, garbage=3, deflate=True)

    # --- 検証 (機械層): 全値の存在 (NFKC + dash 正規化) + ## 残存 + ページ数 ---
    t = flat(fitz.open(out_filled)[0].get_text())
    expected = [str(it["text"]).replace("\n", "") for it in items
                if it.get("verify", True) and it.get("type") != "check"]
    expected += [str(x) for x in assert_present]   # = 雛形 prefill 済の sanity-check (欠落で fail)
    missing = [e for e in expected if flat(e).replace(" ", "") not in t.replace(" ", "").replace("\n", "")]
    if missing or "##" in t or fitz.open(out_filled).page_count != 1:
        raise AssertionError(f"検証 FAIL ({out_base}): missing={missing} hash={'##' in t}")

    # --- 印刷用ラスタ (WYSIWYG 保証) ---
    pg = fitz.open(out_filled)[0]
    png = out_filled.with_suffix(".tmp.png")
    pg.get_pixmap(dpi=dpi).save(png)
    rast = fitz.open()
    np_ = rast.new_page(width=pg.rect.width, height=pg.rect.height)
    np_.insert_image(np_.rect, filename=str(png))
    out_raster = Path(f"{out_base}_raster.pdf")
    rast.save(out_raster, deflate=True)
    png.unlink()
    return {"filled": out_filled, "raster": out_raster}


def verify_pdf_mutation(orig_path, out_path, *,
                        must_present=(), must_absent=(),
                        expected_image_delta: int = 0,
                        text_must_be_identical: bool = True,
                        page_count_must_match: bool = True) -> list:
    """PDF mutation 前後で機械検証 schema 5 項目を回す
    (= [`pdf-mutation-verification-schema`](office-automation.md#pdf-mutation-verification-schema)
    の engine helper 化)。

    1. page 数不変 (image insert / value overlay は page 数を変えない想定)
    2. text content 完全不変 (image-only mutation なら text 1 byte も変わらない想定。
       value overlay や redact なら text_must_be_identical=False、 期待差分を must_present
       で網羅)
    3. MUST-PRESENT keywords が出力に在る (= 雛形 prefill / 過去 hallucination 訂正後の
       正しい値が残っているか)
    4. MUST-ABSENT keywords が出力に無い (= 訂正版 PDF を mutate する時の preservation
       確認、 「視覚は『在って当然のもの』 に目が行き『あるはずの無いもの』 を読み飛ばす」
       認知盲点への machine 点呼)
    5. image stream count delta が想定通り (= 1 枚 insert なら +1)

    text comparison: 全 page を join した raw string で == 比較 (= 単票単独で投げる
    convention。 multi-page form は呼出し側で per-page に invoke 推奨)。

    Returns: list of error string (空 = 全 pass)。 caller は `if errors: sys.exit(1)`
             で送付・印刷 step に進ませない。

    Raises: 引数の型/値が明らかに不正な場合のみ ValueError。 mutation 自体の異常は
            list の error として返す (= 中断せず 5 項目を全て列挙して driver 側で一括処理)。

    See: office-automation.md#pdf-mutation-verification-schema
    """
    if expected_image_delta < 0:
        raise ValueError(f"expected_image_delta must be >= 0, got {expected_image_delta}")
    errors = []
    orig = fitz.open(str(orig_path))
    out = fitz.open(str(out_path))
    try:
        if page_count_must_match and orig.page_count != out.page_count:
            errors.append(f"page 数差: orig={orig.page_count} out={out.page_count}")

        orig_text = "".join(p.get_text() for p in orig)
        out_text = "".join(p.get_text() for p in out)
        if text_must_be_identical and orig_text != out_text:
            for i, (a, b) in enumerate(zip(orig_text, out_text)):
                if a != b:
                    errors.append(
                        f"text 不変性 fail: pos {i}, "
                        f"orig={orig_text[i:i+20]!r}, out={out_text[i:i+20]!r}"
                    )
                    break
            else:
                errors.append(
                    f"text 長さ差: orig={len(orig_text)}, out={len(out_text)}"
                )

        for kw in must_present:
            if kw not in out_text:
                errors.append(f"MUST-PRESENT 不在: {kw!r}")
        for kw in must_absent:
            if kw in out_text:
                errors.append(f"MUST-ABSENT 残存: {kw!r}")

        orig_imgs = sum(len(p.get_images(full=True)) for p in orig)
        out_imgs = sum(len(p.get_images(full=True)) for p in out)
        img_delta = out_imgs - orig_imgs
        if img_delta != expected_image_delta:
            errors.append(
                f"image stream delta 想定外: got={img_delta} expected={expected_image_delta}"
            )
    finally:
        orig.close()
        out.close()
    return errors


def _selftest():
    """programmatic な minimal PDF fixture を fitz で組んで 3 helper を検証 (= 単独実行
    で全 assertion pass 確認)。 PIL/numpy が無ければ A-2 部分を skip。
    実行: `python3 pdf_form_fill.py --selftest`
    """
    import tempfile
    import os
    print("=== pdf_form_fill --selftest ===")

    # ⚠️ selftest fixture は ASCII label のみ使う (= fitz の built-in helv font は CJK
    # glyph を持たず、 CJK を insert_text しても text 層に乗らない / search_for で hit
    # しない。 algorithm 検証は ASCII で十分、 実 PDF (推薦書様式等) での probe は
    # office-automation.md §pdf-cell-label-vs-data-disambiguation origin が cover)
    # ------------------------------------------------------------------
    # Fixture A: horizontal 2 cell (= filled rect approach for tier 1)
    # ------------------------------------------------------------------
    doc = fitz.open()
    page = doc.new_page(width=400, height=200)
    # left filled cell with "Name" text
    page.draw_rect(fitz.Rect(50, 50, 150, 100), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    # right filled cell empty
    page.draw_rect(fitz.Rect(150, 50, 300, 100), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    # label text in left cell
    page.insert_text((70, 80), "Name", fontname="helv", fontsize=10)
    fix_a = Path(tempfile.mkstemp(suffix=".pdf", prefix="fix_A_")[1])
    doc.save(fix_a)
    doc.close()

    da = fitz.open(fix_a)
    r = find_data_cell_for_label(da[0], "Name", direction="right")
    assert abs(r.x0 - 150) < 3 and abs(r.x1 - 300) < 3, f"A right cell wrong: {r}"
    print(f"  ✓ A horizontal: data cell = {r}")
    # default direction=None should also find right (priority order)
    r2 = find_data_cell_for_label(da[0], "Name")
    assert r2 == r, "default direction should find right cell"
    print(f"  ✓ A direction=None: same result")
    # label 不在 → LookupError
    try:
        find_data_cell_for_label(da[0], "NotPresent")
        assert False, "should LookupError"
    except LookupError as e:
        assert "見つからない" in str(e)
        print(f"  ✓ A label not found → LookupError")
    da.close()
    fix_a.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # Fixture B: vertical 2 cell (= filled rect approach, direction=below)
    # ------------------------------------------------------------------
    doc = fitz.open()
    page = doc.new_page(width=400, height=200)
    page.draw_rect(fitz.Rect(50, 30, 200, 70), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    page.draw_rect(fitz.Rect(50, 70, 200, 130), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    page.insert_text((70, 55), "Addr", fontname="helv", fontsize=10)
    fix_b = Path(tempfile.mkstemp(suffix=".pdf", prefix="fix_B_")[1])
    doc.save(fix_b)
    doc.close()

    db = fitz.open(fix_b)
    r = find_data_cell_for_label(db[0], "Addr", direction="below")
    assert abs(r.y0 - 70) < 3 and abs(r.y1 - 130) < 3, f"B below cell wrong: {r}"
    print(f"  ✓ B vertical: data cell = {r}")
    db.close()
    fix_b.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # Fixture C: nested cell (label cell inside outer block) + border-segment data cell
    # (= real-world PDF pattern mimick: outer cell contains label cell, data cell
    # exists only as 4 thin border line segments to the right of label cell)
    # ------------------------------------------------------------------
    doc = fitz.open()
    page = doc.new_page(width=600, height=300)
    # outer block (= label cell, contains label area)
    page.draw_rect(fitz.Rect(100, 50, 250, 150), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    # inner label-tight cell (smaller, contains label text)
    page.draw_rect(fitz.Rect(110, 70, 240, 100), color=(0, 0, 0), fill=(0.95, 0.95, 0.95))
    page.insert_text((130, 90), "Sign", fontname="helv", fontsize=10)
    # data cell drawn as 4 thin border line segments (no filled rect)
    page.draw_line(fitz.Point(250, 50), fitz.Point(450, 50))    # top
    page.draw_line(fitz.Point(250, 150), fitz.Point(450, 150))  # bottom
    page.draw_line(fitz.Point(250, 50), fitz.Point(250, 150))   # left (shared with outer)
    page.draw_line(fitz.Point(450, 50), fitz.Point(450, 150))   # right
    fix_c = Path(tempfile.mkstemp(suffix=".pdf", prefix="fix_C_")[1])
    doc.save(fix_c)
    doc.close()

    dc = fitz.open(fix_c)
    r = find_data_cell_for_label(dc[0], "Sign", direction="right")
    # should find border-segment-constructed cell
    assert r.x0 >= 245 and r.x1 >= 445, f"C border-segment cell wrong: {r}"
    assert dc[0].get_text(clip=r).strip() == "", f"C cell not empty: {dc[0].get_text(clip=r)!r}"
    print(f"  ✓ C nested + border-segment: data cell = {r}")
    dc.close()
    fix_c.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # Fixture D: image-only PDF (= no drawings) → LookupError
    # ------------------------------------------------------------------
    doc = fitz.open()
    page = doc.new_page(width=400, height=200)
    page.insert_text((100, 100), "Lonely", fontname="helv", fontsize=10)
    fix_d = Path(tempfile.mkstemp(suffix=".pdf", prefix="fix_D_")[1])
    doc.save(fix_d)
    doc.close()

    dd = fitz.open(fix_d)
    try:
        find_data_cell_for_label(dd[0], "Lonely")
        assert False, "should LookupError (no drawings)"
    except LookupError as e:
        assert "contain" in str(e) or "drawings" in str(e), f"unexpected msg: {e}"
        print(f"  ✓ D image-only PDF (containing rect 無し) → LookupError")
    dd.close()
    fix_d.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # A-2: boost_signature_alpha (PIL optional)
    # ------------------------------------------------------------------
    try:
        from PIL import Image
        import numpy as np
        has_pil = True
    except ImportError:
        has_pil = False
        print("  ~ A-2 skipped (PIL/numpy not installed)")

    if has_pil:
        from PIL import Image
        import numpy as np
        # gradient alpha PNG fixture
        a = np.linspace(0, 100, 100, dtype=np.uint8)
        a_img = np.tile(a, (50, 1))
        rgba = np.zeros((50, 100, 4), dtype=np.uint8)
        rgba[..., 0] = 128  # mid-gray RGB
        rgba[..., 1] = 128
        rgba[..., 2] = 128
        rgba[..., 3] = a_img
        src = Path(tempfile.mkstemp(suffix=".png", prefix="sig_src_")[1])
        Image.fromarray(rgba).save(src)

        # boost ×3, force_black=True
        tmp = boost_signature_alpha(src, alpha_boost=3.0, force_black=True)
        arr = np.array(Image.open(tmp).convert("RGBA"))
        # alpha should be clipped at 255 for source alpha >= 85
        assert (arr[..., 3].max() == 255), f"alpha max {arr[..., 3].max()} != 255"
        # RGB pixels with alpha>0 should be (0,0,0)
        mask = arr[..., 3] > 0
        if mask.any():
            assert (arr[..., :3][mask] == 0).all(), "force_black violated"
        print(f"  ✓ A-2 boost ×3 + force_black: alpha clipped, RGB pure black")
        tmp.unlink(missing_ok=True)

        # alpha_boost < 0 → ValueError
        try:
            boost_signature_alpha(src, alpha_boost=-1.0)
            assert False, "should ValueError"
        except ValueError as e:
            assert "alpha_boost" in str(e)
            print(f"  ✓ A-2 alpha_boost<0 → ValueError")

        src.unlink(missing_ok=True)

    # ------------------------------------------------------------------
    # A-3: verify_pdf_mutation
    # ------------------------------------------------------------------
    doc1 = fitz.open()
    p1 = doc1.new_page(width=400, height=200)
    p1.insert_text((50, 50), "Hello correct value", fontname="helv", fontsize=10)
    p1.insert_text((50, 80), "B prior correction", fontname="helv", fontsize=10)
    orig_p = Path(tempfile.mkstemp(suffix=".pdf", prefix="orig_")[1])
    doc1.save(orig_p)
    doc1.close()

    # identical copy (image-only mutation simulation)
    import shutil
    same_p = Path(tempfile.mkstemp(suffix=".pdf", prefix="same_")[1])
    shutil.copy(orig_p, same_p)

    # case 1: same PDF, no image delta → 全 pass
    errs = verify_pdf_mutation(orig_p, same_p,
                               must_present=["Hello", "correct value"],
                               must_absent=["XYZ hallucination"],
                               expected_image_delta=0)
    assert errs == [], f"case1 should pass: {errs}"
    print(f"  ✓ A-3 identical PDF: 0 errors")

    # case 2: text differs but text_must_be_identical=True → fail
    doc2 = fitz.open()
    p2 = doc2.new_page(width=400, height=200)
    p2.insert_text((50, 50), "Hello DIFFERENT", fontname="helv", fontsize=10)
    diff_p = Path(tempfile.mkstemp(suffix=".pdf", prefix="diff_")[1])
    doc2.save(diff_p)
    doc2.close()
    errs = verify_pdf_mutation(orig_p, diff_p,
                               must_present=[], must_absent=[],
                               expected_image_delta=0)
    assert any("text" in e for e in errs), f"case2 should have text error: {errs}"
    print(f"  ✓ A-3 text differs: caught ({len(errs)} errors)")

    # case 3: MUST-PRESENT missing → fail
    errs = verify_pdf_mutation(orig_p, same_p,
                               must_present=["完全に不在の語"],
                               must_absent=[], expected_image_delta=0)
    assert any("MUST-PRESENT" in e for e in errs), f"case3 missing detection: {errs}"
    print(f"  ✓ A-3 MUST-PRESENT missing: caught")

    # case 4: MUST-ABSENT residual → fail
    errs = verify_pdf_mutation(orig_p, same_p,
                               must_present=[], must_absent=["Hello"],
                               expected_image_delta=0)
    assert any("MUST-ABSENT" in e for e in errs), f"case4 residual detection: {errs}"
    print(f"  ✓ A-3 MUST-ABSENT residual: caught")

    # case 5: image delta mismatch → fail
    errs = verify_pdf_mutation(orig_p, same_p,
                               must_present=[], must_absent=[],
                               expected_image_delta=1)
    assert any("image stream" in e for e in errs), f"case5 image delta: {errs}"
    print(f"  ✓ A-3 image delta mismatch: caught")

    # case 6: text_must_be_identical=False allows text drift
    errs = verify_pdf_mutation(orig_p, diff_p,
                               must_present=[], must_absent=[],
                               text_must_be_identical=False,
                               expected_image_delta=0)
    text_errs = [e for e in errs if "text" in e]
    assert text_errs == [], f"case6 text drift should be allowed: {text_errs}"
    print(f"  ✓ A-3 text_must_be_identical=False: text drift allowed")

    # case 7: expected_image_delta<0 → ValueError
    try:
        verify_pdf_mutation(orig_p, same_p, expected_image_delta=-1)
        assert False, "should ValueError"
    except ValueError as e:
        assert "expected_image_delta" in str(e)
        print(f"  ✓ A-3 expected_image_delta<0 → ValueError")

    orig_p.unlink(missing_ok=True)
    same_p.unlink(missing_ok=True)
    diff_p.unlink(missing_ok=True)

    print("=== ALL PASS ===")


if __name__ == "__main__":
    import sys
    if "--selftest" in sys.argv:
        _selftest()
    else:
        print("Usage: python3 pdf_form_fill.py --selftest")
        sys.exit(2)
