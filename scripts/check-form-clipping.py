#!/usr/bin/env python3
"""
check-form-clipping.py — 生成 form PDF で「記入値が描画時に clip された」のを機械検出。

WHY (office-automation-principles の検証 3 層 = 機械/視覚/実機。 従来その「機械」層は
文字 clipping を *捕捉不能* と設計上あきらめ、 視覚層 (人の目) 頼りだった
[principles.md §2 の表「捕まえられないもの = 文字切れ」]。 人が見落とせば素通り。
2026-06 謝金様式⑭-1 の所属見切れ RCA = 結合セル (例 G13:M13) に長い所属名 (16 字相当)
を固定 14pt で書くと両端が clip → セル値は完全なので diff-form-xlsx では出ず、 視覚確認も
所属行を凝視しないと気づかない → 人の目だけが gate だった)。 ここを埋める。

WHAT: 雛形 (template) と記入済 (filled) を diff して **記入セル (= 値が変わったセル)**
だけを取り、 各記入値が生成 PDF の抽出テキスト (全ページ連結 + NFKC 正規化 + 空白除去)
に「*完全な* 部分文字列として現れるか」を照合する。
  - 完全に現れる                  → OK
  - 一部だけ (= 最長片はあるが 3 字以上が欠ける) → CLIP 疑い (描画で truncate)
  - ほぼ現れない (= 最長片 < 3 字)  → 別ページ/別 sheet (対象外、 flag しない)

雛形 diff で「記入値だけ」に絞るのが肝 (= テンプレの注意書きラベルや他 sheet の固定文を
誤検出しない。 diff-form-xlsx と同じ契約)。 閾値「欠落 ≥ 3 字」で正規化由来の 1-2 字差も除外。

LIMIT (= 視覚確認を置換しない): レンダラによっては clip しても text 層に全文を残す
(clip-path 方式) ことがある。 LibreOffice / Excel は実測で truncate するため本法で捕まるが
engine 依存。 本検出は「機械層の第一防衛線」 であって pdf-visual-confirm を置換しない。

USAGE:
  check-form-clipping.py <template.xlsx> <filled.xlsx> <form.pdf>   # exit 1 if 疑いあり
  check-form-clipping.py --min-len N <tpl> <filled> <pdf>           # 最短照合字数 (既定 4)
  check-form-clipping.py --selftest                                 # 内蔵 self-test (file 不要)
"""

from __future__ import annotations

import argparse
import sys
import unicodedata

_GLYPH_SKIP = set("☑□■✓✔☐−—–-〇○●◯ 　")
MISSING_MIN = 3   # この字数以上が欠けたら clip とみなす (1-2 字は正規化/描画差として無視)
FRAG_MIN = 3      # この字数以上の連続片が在る = そのセルがこの PDF 上に在る証拠


def _norm(s: str) -> str:
    """NFKC 正規化 + 空白除去 (= ⾼→高 等の互換字、 折り返し空白、 全/半角差を吸収)。"""
    s = unicodedata.normalize("NFKC", s)
    return "".join(ch for ch in s if not ch.isspace())


def _longest_common_substring_len(needle: str, hay: str) -> int:
    best = 0
    n = len(needle)
    for i in range(n):
        if n - i <= best:
            break
        j = i + best + 1
        while j <= n and needle[i:j] in hay:
            best = j - i
            j += 1
    return best


def find_clipping(values, pdf_text, min_len=4):
    """純関数 (= selftest 可能)。 values: [(loc, raw_value)] / pdf_text: str。
    返り値: [(loc, value, frag_len, value_len)] = clip 疑いのみ。"""
    hay = _norm(pdf_text)
    flagged = []
    for loc, raw in values:
        if not isinstance(raw, str) or raw.startswith("="):
            continue
        needle = _norm(raw)
        if len(needle) < min_len or all(ch in _GLYPH_SKIP for ch in needle):
            continue
        if needle in hay:
            continue  # 完全に在る = OK
        frag = _longest_common_substring_len(needle, hay)
        if frag >= FRAG_MIN and (len(needle) - frag) >= MISSING_MIN:
            flagged.append((loc, raw, frag, len(needle)))
        # frag 小 = 別ページ/別 sheet / 欠落 1-2 字 = 描画差 → flag しない
    return flagged


def _input_cells(template_path, filled_path):
    """雛形と記入済を diff して、 値が変わった (= 記入された) 文字列セルだけ返す。"""
    import openpyxl
    tpl = openpyxl.load_workbook(template_path, data_only=True)
    fil = openpyxl.load_workbook(filled_path, data_only=True)
    tmap = {}
    for ws in tpl.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    tmap[(ws.title, c.coordinate)] = c.value
    out = []
    for ws in fil.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.strip() and tmap.get((ws.title, c.coordinate)) != v:
                    out.append((f"{ws.title}!{c.coordinate}", v))
    return out


def _extract_pdf_text(path):
    import fitz
    d = fitz.open(path)
    return "\n".join(pg.get_text() for pg in d)


def _selftest():
    # 注: literal は全て架空 (= 公開 layer の PII leak 防止)。 ロジックを exercise するだけ。
    cases = []
    vals = [("S!G13", "国立アイウエオ加速器研究機構（略称）"),
            ("S!Y13", "甲野 太郎"), ("S!Q13", "研究員")]
    pdf = "所属 アイウエオ加速器研究機構（ 職名 研究員 氏名 甲野 太郎"
    f = find_clipping(vals, pdf)
    cases.append(("clip 検出 (結合セルの長文)", [l for l, *_ in f] == ["S!G13"]))
    cases.append(("非clip (短い所属は収まる)", find_clipping(
        [("S!G13", "架空女子大学大学院")], "所属 架空女子大学大学院 職名") == []))
    cases.append(("NFKC 互換字を同一視", find_clipping(
        [("S!A1", "あいう大学")], "あいう⼤学") == []))  # ⼤=U+2F24 → NFKC → 大
    cases.append(("別ページ非検出", find_clipping(
        [("S!A1", "まったく無関係な長い文字列")], "ここには別の短い内容のみ") == []))
    cases.append(("折返し空白吸収", find_clipping(
        [("S!A1", "カキクケコサシスセソ")], "カキクケコ\nサシスセソ") == []))
    cases.append(("欠落1字は無視 (先頭タブ)", find_clipping(
        [("S!A1", "\tあいうえおかきくけこさしすせそ")],
        "あいうえおかきくけこさしすせそ") == []))
    cases.append(("部分一致(欠落2字)は無視", find_clipping(
        [("S!A1", "あい承認欄")], "ここに承認欄あり") == []))  # frag=3 だが欠落2字
    ok = True
    for name, res in cases:
        print(f"  [{'PASS' if res else 'FAIL'}] {name}")
        ok = ok and res
    print("selftest:", "ALL PASS" if ok else "FAILED")
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template", nargs="?")
    ap.add_argument("filled", nargs="?")
    ap.add_argument("pdf", nargs="?")
    ap.add_argument("--min-len", type=int, default=4)
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    if args.selftest:
        sys.exit(_selftest())
    if not (args.template and args.filled and args.pdf):
        ap.error("template.xlsx filled.xlsx form.pdf を指定 (または --selftest)")

    values = _input_cells(args.template, args.filled)
    text = _extract_pdf_text(args.pdf)
    flagged = find_clipping(values, text, min_len=args.min_len)

    if not flagged:
        print(f"✓ clip 疑いなし (記入セル {len(values)} 件を照合)")
        sys.exit(0)
    print(f"⚠️  clip 疑い {len(flagged)} 件 (= 記入値が PDF 描画で truncate された可能性):")
    for loc, val, frag, total in flagged:
        print(f"  {loc}: 値「{val}」({total}字) のうち最長 {frag}字 しか描画されていない")
    print("対処: 結合セルなら font size を下げる (shrink_to_fit は無効)。")
    print("      office-automation.md#merged-cell-text-clipping 参照。視覚確認も併用。")
    sys.exit(1)


if __name__ == "__main__":
    main()
